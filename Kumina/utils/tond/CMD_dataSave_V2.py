from openpyxl import workbook
from openpyxl import load_workbook
from utils.tond.dataUtils import DATA_UTILS
from utils.tond.CMD_dataCheck import COMMAND_DATACHECK
import os


class COMMAND_DATASAVE_V2:
    def __init__(self):
        self.dataUtils = DATA_UTILS()
        self.dataCheck = COMMAND_DATACHECK()

    def runDataSave(self, loc, msg, cellLoc, fileName):
        if os.path.exists(fileName):
            wb = load_workbook(fileName)
        else:
            wb = workbook.Workbook()

        sheet = wb.active

        cellValue = msg

        cell = sheet[cellLoc]

        if cell.value is None:
            cell.value = '{}'

        cell.value = self.dataUtils.saveToJsonString(loc, cellValue, cell.value)

        wb.save(fileName)

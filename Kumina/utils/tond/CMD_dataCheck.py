from openpyxl import load_workbook
from openpyxl import workbook
import os


class COMMAND_DATACHECK:
    def __init__(self):
        self.alphabet = [
            'A',
            'B',
            'C',
            'D',
            'E',
            'F',
            'G',
            'H',
            'I',
            'J',
            'K',
            'L',
            'M',
            'N',
            'O',
            'P',
            'Q',
            'R',
            'S',
            'T',
            'U',
            'V',
            'W',
            'X',
            'Y',
            'Z'
]

    def runCheckData(self):
        tempStore = self.dataCheck()
        print(f'output> Closest Available Storage is [{tempStore}]')

    def dataCheck(self):
        fileName = '_data_.xlsx'

        wb = None

        if os.path.exists(fileName):
            wb = load_workbook(fileName)
        else:
            wb = workbook.Workbook()

        sheet = wb.active

        for row in range(1, sheet.max_column + 2):
            for column in range(1, sheet.max_column + 2):

                tempCell = sheet.cell(row, column)

                if tempCell.value != None:
                    tempStore1 = row
                    tempStore2 = column
                else:
                    tempStore1 = row
                    tempStore2 = column
                    break

        print(tempStore1)
        print(tempStore1 / 26)

        if tempStore1 / 26 < 1:
            tempStore1 = self.alphabet[tempStore1 - 1]
        else:
            tempStoreI = self.alphabet[tempStore1 - 1 - tempStore1 % 26]
            tempStore1 = tempStoreI * round(tempStore1 / 26)
        tempStore = tempStore1 + str(tempStore2)

        return tempStore
